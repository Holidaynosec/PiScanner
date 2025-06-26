import asyncio
import json
import logging
import aiohttp
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from generator.generator_api import APIGenerator
from generator.generator_ollama import OllamaGenerator
from generator.generator_openai import OpenAIGenerator
from evaluator import PromptInjectionEvaluator

class PiScannerHost:
    def __init__(self, config_dir: str = "configuration"):
        """Initialize main controller"""
        self.config_dir = Path(config_dir)
        self.agent_config_path = self.config_dir / "agent_conf.json"
        self.evaluator_config_path = self.config_dir / "evaluator_conf.json"
        self.prompts_file = self.config_dir / "injected_prompts.txt"
        
        self.logger = logging.getLogger(__name__)
        Path("output").mkdir(exist_ok=True)
        self.load_config()
        
    def load_config(self):
        """Load configuration"""
        with open(self.agent_config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
    def generate_output_filename(self, agent_type: str) -> str:
        """Generate timestamped output filename"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return f"output/pi_scaned_{agent_type}_{timestamp}.xlsx"
        
    async def run_generation(self, agent_type: str):
        """Run content generation and evaluation process"""
        # 1. Initialize generator and evaluator
        generator = self._get_generator(agent_type)
        evaluator = PromptInjectionEvaluator(str(self.evaluator_config_path))
        
        # 2. Generate AI responses
        self.logger.info("📝 Starting AI response generation...")
        generation_results = await generator.send_prompts(str(self.prompts_file))
        self.logger.info(f"✅ Response generation completed, {len(generation_results)} results")
        
        # 3. Evaluate results
        self.logger.info("🔍 Starting evaluation aggregation...")
        evaluation_results = await self._evaluate_results(evaluator, generation_results)
        
        # 4. Save results
        self.save_results(evaluation_results, agent_type)
        
        return evaluation_results

    def _get_generator(self, agent_type: str):
        """Get content generator instance"""
        generators = {
            "api": APIGenerator,
            "ollama": OllamaGenerator,
            "openai": OpenAIGenerator
        }
        
        if agent_type not in generators:
            raise ValueError(f"Unsupported agent type: {agent_type}")
            
        return generators[agent_type](str(self.agent_config_path))

    async def _evaluate_results(self, evaluator: PromptInjectionEvaluator, generation_results: List[Dict]) -> List[Dict]:
        """Evaluate generation results"""
        evaluation_results = []
        
        for i, result in enumerate(generation_results):
            if result['status'] == 'success':
                evaluation = await evaluator.evaluate_pair(result['prompt'], result['response'])
            else:
                evaluation = self._create_failed_evaluation(result)
                self.logger.info(f"Error: {evaluation['response']}")
            
            evaluation_results.append(evaluation)
            self.logger.info(f"📊 Evaluation progress: {i+1}/{len(generation_results)}")
        
        return evaluation_results

    def _create_failed_evaluation(self, result: Dict) -> Dict:
        """Create evaluation result for failed requests"""
        return {
            'prompt': result['prompt'],
            'response': result['response'],
            'injected_result': False,
            'llm_evaluation_reason': f"Request failed: {result['status']}",
            'compliance_score': 0.0
        }
        
    def save_results(self, results: List[Dict[str, Any]], agent_type: str):
        """Save results to Excel file"""
        output_file = self.generate_output_filename(agent_type)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"PI_Scan_{agent_type}"
        
        # Define headers and styles
        headers = ['tested_prompt', 'AI_response', 'injected_result', 'llm_evaluation_reason', 'compliance_score']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result['prompt'])
            ws.cell(row=row_idx, column=2, value=result['response'])
            ws.cell(row=row_idx, column=3, value=result['injected_result'])
            ws.cell(row=row_idx, column=4, value=result['llm_evaluation_reason'])
            ws.cell(row=row_idx, column=5, value=result['compliance_score'])
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = min(max(max_length + 2, 15), 100)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(output_file)
        wb.close()
        
        self.logger.info(f"💾 Results saved to: {output_file}")
        self.print_summary(results, agent_type)
        
    def print_summary(self, results: List[Dict[str, Any]], agent_type: str):
        """Print generation and evaluation summary"""
        total = len(results)
        successful_injections = sum(1 for r in results if r['injected_result'] == 'success')
        
        self.logger.info("=" * 50)
        self.logger.info(f"📊 Generation and Evaluation Summary - Target Type: {agent_type.upper()}")
        self.logger.info("=" * 50)
        self.logger.info(f"Total test samples: {total}")
        self.logger.info(f"Successful injections: {successful_injections}")
        self.logger.info(f"Injection success rate: {successful_injections/total*100:.1f}%")
        self.logger.info("=" * 50)

    async def run_all_enabled_generations(self):
        """Run all enabled content generation"""
        tasks = []
        agent_configs = {
            'api': self.config['agent_types'].get('api_agent', {}),
            'ollama': self.config['agent_types'].get('ollama_agent', {}),
            'openai': self.config['agent_types'].get('openai_agent', {})
        }
        
        # Add enabled tasks
        for agent_type, config in agent_configs.items():
            if config.get('enabled', False):
                tasks.append(self.run_generation(agent_type))
        
        if tasks:
            return await asyncio.gather(*tasks, return_exceptions=True)
        else:
            self.logger.warning("⚠️ No enabled agent types")
            return []

    async def run_specific_generation(self, agent_type: str):
        """Run specific type of content generation"""
        if agent_type not in ['api', 'ollama', 'openai']:
            self.logger.error(f"❌ Unsupported agent type: {agent_type}")
            return None
            
        return await self.run_generation(agent_type)

    async def run_single_prompt_generation(self, agent_type: str, prompt: str):
        """Run single prompt generation and evaluation process"""
        if agent_type not in ['api', 'ollama', 'openai']:
            self.logger.error(f"❌ Unsupported agent type: {agent_type}")
            return None
        
        # 1. Initialize generator and evaluator
        generator = self._get_generator(agent_type)
        evaluator = PromptInjectionEvaluator(str(self.evaluator_config_path))
        
        # 2. Generate AI response
        self.logger.info("📝 Starting AI response generation...")
        
        if agent_type == "api":
            async with aiohttp.ClientSession() as session:
                generation_result = await generator.send_request(session, prompt)
        elif agent_type == "ollama":
            async with aiohttp.ClientSession() as session:
                generation_result = await generator.send_request(session, prompt)
        else:  # openai
            generation_result = generator.send_request(prompt)
        
        self.logger.info("✅ Response generation completed")
        
        # 3. Evaluate results
        self.logger.info("🔍 Starting result evaluation...")
        evaluation_results = await self._evaluate_results(evaluator, [generation_result])
        
        # 4. Print results
        self.print_single_result(evaluation_results[0])
        
        return evaluation_results

    def print_single_result(self, result: Dict[str, Any]):
        """Print key information of single result"""
        print("=" * 60)
        self.logger.info("📊 Detection Results")
        self.logger.info(f"📝 Prompt: {result['prompt']}")
        self.logger.info(f"🎯 Injection Result: {result['injected_result']}")
        self.logger.info(f"📋 Evaluation Reason: {result['llm_evaluation_reason']}")
        print("=" * 60)
